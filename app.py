import os, json, io, sqlite3
from datetime import datetime
from flask import Flask, request, jsonify, render_template, send_file, session, redirect, url_for
import anthropic
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "reventa-secret-2024")

ANTHROPIC_API_KEY = os.environ.get("ANTHROPIC_API_KEY", "")
LOGIN_PASSWORD    = os.environ.get("LOGIN_PASSWORD", "reventa123")
DB_PATH           = os.environ.get("DB_PATH", "pedidos.db")

# ── Base de datos ─────────────────────────────────────────────────────────────
def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
      conn = sqlite3.connect(DB_PATH)
      conn.execute("""
          CREATE TABLE IF NOT EXISTS pedidos (
              id            INTEGER PRIMARY KEY AUTOINCREMENT,
              fecha         TEXT NOT NULL,
              cliente       TEXT NOT NULL,
              producto      TEXT NOT NULL,
              cantidad      REAL NOT NULL DEFAULT 1,
              precio_venta  REAL NOT NULL DEFAULT 0,
              precio_costo  REAL NOT NULL DEFAULT 0,
              total_venta   REAL NOT NULL DEFAULT 0,
              total_costo   REAL NOT NULL DEFAULT 0,
              ganancia      REAL NOT NULL DEFAULT 0,
              margen        REAL NOT NULL DEFAULT 0,
              notas         TEXT DEFAULT ''
          )
      """)
      conn.commit()
      conn.close()
# ── Auth ──────────────────────────────────────────────────────────────────────
def logged_in():
    return session.get("auth") is True

@app.route("/login", methods=["GET", "POST"])
def login():
    error = None
    if request.method == "POST":
        if request.form.get("password") == LOGIN_PASSWORD:
            session["auth"] = True
            return redirect(url_for("index"))
        error = "Contraseña incorrecta"
    return render_template("login.html", error=error)

@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))

# ── Páginas ───────────────────────────────────────────────────────────────────
@app.route("/")
def index():
    if not logged_in():
        return redirect(url_for("login"))
    return render_template("index.html")

# ── API: extracción con Claude ────────────────────────────────────────────────
PROMPT = """Sos asistente de una reventa. Del siguiente texto extraé los datos del pedido en JSON.
Devolvé SOLO el JSON, sin texto extra ni bloques de código.

Formato:
{{
  "cliente": "nombre del cliente",
  "productos": [
    {{
      "nombre": "nombre del producto",
      "cantidad": 2,
      "precio_venta": 500,
      "precio_costo": 350
    }}
  ],
  "notas": "observación o null"
}}

Si un dato no se menciona, usá null. cantidad debe ser un número.

Texto: {texto}"""

@app.route("/api/extract", methods=["POST"])
def extract():
    if not logged_in():
        return jsonify({"error": "No autorizado"}), 401
    texto = (request.json or {}).get("texto", "").strip()
    if not texto:
        return jsonify({"error": "Texto vacío"}), 400
    client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)
    resp = client.messages.create(
        model="claude-sonnet-4-6",
        max_tokens=1024,
        messages=[{"role": "user", "content": PROMPT.format(texto=texto)}]
    )
    raw = resp.content[0].text.strip()
    if raw.startswith("```"):
        parts = raw.split("```")
        raw = parts[1] if len(parts) > 1 else raw
        if raw.startswith("json"):
            raw = raw[4:]
    return jsonify(json.loads(raw.strip()))

# ── API: guardar pedido ───────────────────────────────────────────────────────
@app.route("/api/orders", methods=["POST"])
def save_order():
    if not logged_in():
        return jsonify({"error": "No autorizado"}), 401
    data    = request.json or {}
    fecha   = datetime.now().strftime("%d/%m/%Y %H:%M")
    cliente = (data.get("cliente") or "Sin nombre").strip()
    notas   = data.get("notas") or ""
    items   = []
    with get_db() as db:
        for prod in data.get("productos") or []:
            nombre   = (prod.get("nombre") or "Producto").strip()
            cantidad = float(prod.get("cantidad") or 1)
            pv       = float(prod.get("precio_venta") or 0)
            pc       = float(prod.get("precio_costo") or 0)
            tv       = round(cantidad * pv, 2)
            tc       = round(cantidad * pc, 2)
            gan      = round(tv - tc, 2)
            margen   = round(gan / tv * 100, 1) if tv > 0 else 0
            db.execute(
                "INSERT INTO pedidos VALUES (NULL,?,?,?,?,?,?,?,?,?,?,?)",
                (fecha, cliente, nombre, cantidad, pv, pc, tv, tc, gan, margen, notas)
            )
            items.append({"producto": nombre, "cantidad": cantidad,
                          "total_venta": tv, "margen": margen})
        db.commit()
    return jsonify({"ok": True, "cliente": cliente, "items": items})

# ── API: listar pedidos ───────────────────────────────────────────────────────
@app.route("/api/orders", methods=["GET"])
def get_orders():
    if not logged_in():
        return jsonify({"error": "No autorizado"}), 401
    cliente = request.args.get("cliente")
    with get_db() as db:
        if cliente:
            rows = db.execute(
                "SELECT * FROM pedidos WHERE cliente=? ORDER BY id DESC LIMIT 50",
                (cliente,)
            ).fetchall()
        else:
            rows = db.execute(
                "SELECT * FROM pedidos ORDER BY id DESC LIMIT 50"
            ).fetchall()
    return jsonify([dict(r) for r in rows])

# ── API: clientes ─────────────────────────────────────────────────────────────
@app.route("/api/clients")
def get_clients():
    if not logged_in():
        return jsonify({"error": "No autorizado"}), 401
    with get_db() as db:
        rows = db.execute(
            "SELECT cliente, COUNT(*) as pedidos, SUM(total_venta) as total "
            "FROM pedidos GROUP BY cliente ORDER BY cliente"
        ).fetchall()
    return jsonify([dict(r) for r in rows])

# ── API: exportar Excel ───────────────────────────────────────────────────────
COLS = ["Fecha","Cliente","Producto","Cantidad","P. Venta Unit.","P. Costo Unit.",
        "Total Venta","Total Costo","Ganancia","Margen %","Notas"]
WIDTHS = [16, 22, 22, 10, 15, 15, 13, 13, 13, 10, 28]

@app.route("/api/export")
def export():
    if not logged_in():
        return redirect(url_for("login"))
    cliente = request.args.get("cliente")
    with get_db() as db:
        if cliente:
            rows = db.execute("SELECT * FROM pedidos WHERE cliente=? ORDER BY id",
                              (cliente,)).fetchall()
            fname = f"pedidos_{cliente.lower().replace(' ','_')}.xlsx"
        else:
            rows = db.execute("SELECT * FROM pedidos ORDER BY id").fetchall()
            fname = "pedidos_todos.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pedidos"
    fill = PatternFill("solid", fgColor="1F4E79")
    hfont = Font(color="FFFFFF", bold=True)
    for col, (h, w) in enumerate(zip(COLS, WIDTHS), 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = fill; c.font = hfont
        c.alignment = Alignment(horizontal="center")
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w
    for r in rows:
        r = dict(r)
        ws.append([r["fecha"], r["cliente"], r["producto"], r["cantidad"],
                   r["precio_venta"], r["precio_costo"], r["total_venta"],
                   r["total_costo"], r["ganancia"], r["margen"], r.get("notas","")])

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ── Main ──────────────────────────────────────────────────────────────────────
init_db()

@app.before_request
def ensure_db():
    init_db()

if __name__ == "__main__":
    app.run(debug=True, host="0.0.0.0", port=5000)
